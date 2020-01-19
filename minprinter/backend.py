#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" The backend functions.
Author: sccotte@gmail.com
"""

import itertools
import logging
import operator
import os
import platform
import re
from copy import copy
from glob import glob
from os.path import basename, join
from subprocess import PIPE, Popen
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from pdf2image import convert_from_path, pdfinfo_from_path
from PIL import Image

logger = logging.getLogger('Backend')
logger.setLevel(logging.DEBUG)


class MinvoiceException(Exception):
    """Exception class for all backend functions.
    """
    def __init__(self, message):
        """Constructor"""
        self.message = message

    def __str__(self):
        """Reprentation of the instance"""
        return repr(self.message)


def find_pdfs(pdfs_dir, recursive=True, exclude_basenames=None):
    """ Find the invoice PDF file from the specified directory. The invoice
        PDF file should contains only one page.
    """
    if not pdfs_dir.endswith('/'):
        pdfs_dir += '/'
    # for windows, iglob is case-insensitive
    if recursive is True:
        pdfs = glob(pdfs_dir + '**/*.pdf', recursive=True)
    else:
        pdfs = glob(pdfs_dir + '*.pdf', recursive=False)
    if not pdfs:
        raise MinvoiceException('No invoice PDF file found!')
    pdfs_return = []
    for pdf_path in pdfs:
        if exclude_basenames is not None and \
                basename(pdf_path) in exclude_basenames:
            continue
        page_count = pdfinfo_from_path(pdf_path)["Pages"]
        if page_count != 1:
            msg = """Fatal error in processing PDF file {}: the invoice should
be one page per file""".format(pdf_path)
            raise MinvoiceException(msg)
        pdfs_return.append(pdf_path)
    return pdfs_return


def _get_command_path(command, poppler_path=None):
    """Return poppler command path from program name.
    """
    if platform.system() == "Windows":
        command = command + ".exe"
    if poppler_path is not None:
        command = os.path.join(poppler_path, command)
    return command


def pdftotext_from_path(pdf_path, text_path, poppler_path=None):
    """ Extract all pdf text to a txt file.
        The external program `pdftotext` is used, with option -layout
    """
    try:
        command = [_get_command_path("pdftotext", poppler_path), '-layout']

        command.extend([pdf_path, text_path])
        # Add poppler path to LD_LIBRARY_PATH
        env = os.environ.copy()
        if poppler_path is not None:
            env["LD_LIBRARY_PATH"] = poppler_path + ":" + env.get(
                "LD_LIBRARY_PATH", "")
        proc = Popen(command, env=env, stdout=PIPE, stderr=PIPE)
        out, err = proc.communicate()
        logger.debug('out=%s, error=%s', out, err)
    except (OSError, ValueError) as err:
        logger.error(err)
        raise


def to_text_str(pdf_path, poppler_path=None):
    """Extract text from a PDF file and return it as string.
    """
    ntp = NamedTemporaryFile('w+t', suffix='.txt', delete=False)
    ntp.close()
    pdftotext_from_path(pdf_path, ntp.name, poppler_path=poppler_path)
    with open(ntp.name, 'rt', encoding='utf-8') as fp:
        text_str = fp.read()
    os.unlink(ntp.name)
    return text_str


def parse_text(text_str):
    """ Parse the text content from invoice PDF to get name, phone and billing
        amount from a user.
    """
    user_name_regex = r'名\s*称\W\s?([\u4e00-\u9fff]{2,})\s+'
    user_phoneno_regex = r'号码\W\s?(\d{11})'
    bill_date_regex = r'[账帐]期\W\s?(\d{6})'
    bill_amount_regex = r'\W小写\W+(\d+\.\d{2})'
    r1 = re.search(user_name_regex, text_str, flags=re.U)
    r2 = re.search(user_phoneno_regex, text_str, flags=re.U)
    r3 = re.search(bill_date_regex, text_str, flags=re.U)
    r4 = re.search(bill_amount_regex, text_str, flags=re.U)
    any_fail_flag = False
    if r1:
        user_name = r1.group(1)
        if len(user_name) >= 4:
            logger.warning('Name %r is too long, use the last three chars.',
                           user_name)
            user_name = user_name[-3:]
    else:
        any_fail_flag = True
        logger.error('Cannot extract user name with regular express %r',
                     user_name_regex)
    if r2:
        user_phoneno = r2.group(1)
    else:
        any_fail_flag = True
        logger.error(
            'Cannot extract user phone number with regular express %r',
            user_phoneno_regex)
    if r3:
        bill_date = r3.group(1)
    else:
        any_fail_flag = True
        logger.error(
            'Cannot extract user billing date with regular express %r',
            bill_date_regex)
    if r4:
        bill_amount = r4.group(1)
    else:
        any_fail_flag = True
        logger.error(
            'Cannot extract user billing amount with regular express %r',
            bill_amount_regex)
    if any_fail_flag is False:
        return (user_name, user_phoneno, bill_date, bill_amount)
    raise MinvoiceException(
        'Parse invoice text failed: {}, check logs'.format(text_str))


def pdf_to_jpg(pdfs, dpi=600):
    """Coverte pdf to jpe image file with specified dpi"""
    images = []
    for pdf_path in pdfs:
        # to image
        images += convert_from_path(pdf_path, dpi=dpi)
    return images


def to_raw_jpg_pdf(images, outfile):
    """ Save jpg images as a PDF file, each jpg image as one pdf page.
    """
    images[0].save(outfile, "PDF", save_all=True, append_images=images[1:])


def box_from_a4(image_size, a4_size, top_or_bottom=True):
    """Get the box when paste image into a4 sized page. If image size is
    overflow, then a resize operation on longer side would be taken while
    the with/height ratio is kept.

    Parameters:
        image_size: tuple
          (image_width, image_height)
        a4_size: tuple
          (a4_width, a4_height)
        top_or_bottom: boolean
          True: image is placed at top half of the a4 page
          False: image is place at bottom half of the a4 page
    """
    im_w, im_h = image_size
    a4_w, a4_h = a4_size

    w_r = im_w / a4_w
    h_r = im_h / a4_h * 2
    im_re_size = None
    # image overflow the a4, resize is needed
    if w_r > 1 or h_r > 1:
        # width is much larger
        if w_r > h_r:
            im_re_w, im_re_h = (a4_w, int(im_h / w_r))
            if top_or_bottom is True:
                box_top_left = (0, int((a4_h - im_re_h * 2) / 4))
            else:
                box_top_left = (0, int((a4_h - im_re_h * 2) / 2) + im_re_h)
        # height is much larger
        else:
            im_re_w, im_re_h = (int(im_w / h_r), int(a4_h / 2))
            if top_or_bottom is True:
                box_top_left = (int((a4_w - im_re_w) / 2), 0)
            else:
                box_top_left = (int((a4_w - im_re_w) / 2), im_re_h)
        im_re_size = (im_re_w, im_re_h)
    else:
        if top_or_bottom is True:
            box_top_left = (int((a4_w - im_w) / 2), int((a4_h - im_h) / 4))
        else:
            box_top_left = (int((a4_w - im_w) / 2), int(
                (a4_h - im_h) / 2) + im_h)
    return box_top_left, im_re_size


def place_one_jpg_on_a4_page(image, a4_page, top_or_bottom):
    """Place one jpg image on the A4 sized page.
    """
    box_top_left, im_re_size = box_from_a4(image.size,
                                           a4_page.size,
                                           top_or_bottom=top_or_bottom)
    if im_re_size is not None:
        resized_image = image.resize(im_re_size)
        a4_page.paste(resized_image, box=box_top_left)
    else:
        a4_page.paste(image, box=box_top_left)


def to_a4_jpg_pdf(images, outfile, dpi=600):
    """ Save jpg images as a A4-sized PDF file, two jpg images are placed
    in the upper and lower center of the A4 paper. Resize operation could
    be taken.
    """
    # A4 size in pixel
    a4_size = int(8.27 * dpi), int(11.7 * dpi)
    pages = []
    for i in range(0, len(images), 2):
        page = Image.new('RGB', a4_size, 'white')
        place_one_jpg_on_a4_page(images[i], page, top_or_bottom=True)
        if (i + 1) < len(images):
            place_one_jpg_on_a4_page(images[i + 1], page, top_or_bottom=False)
        pages.append(page)
    pages[0].save(outfile, "PDF", save_all=True, append_images=pages[1:])


def year_and_quarter(date_str):
    """Return the year and quarter as numeric.
    date_str should be 'yyyymm'
    """
    date_str = str(date_str)
    year = int(date_str[:4])
    month = int(date_str[4:6])
    quarter = (month - 1) // 3 + 1
    return (year, quarter)


def groupby(l, by):
    """return k, group: where k is the unique key and group is an iterable.

    Paramters
    -----------
      l: list of list
      by: list of integer
        the column indexes
    """
    ll = sorted(l, key=operator.itemgetter(*by))
    return itertools.groupby(ll, key=operator.itemgetter(*by))


def excel_merge_cells(ws,
                      start_row,
                      start_column,
                      end_row,
                      end_column,
                      value=None):
    """A quick function to merge excel cells and set alignment as centered.
    """
    ws.merge_cells(start_row=start_row,
                   end_row=end_row,
                   start_column=start_column,
                   end_column=end_column)
    lt_cell = ws.cell(row=start_row, column=start_column)
    if value is not None:
        lt_cell.value = value


def set_outline_border(ws,
                       border_style,
                       min_row,
                       max_row,
                       min_col=1,
                       max_col=7,
                       color='000000'):
    """Set the outline border with box (min_row, max_row, min_col, max_col).
    """
    border = Side(border_style=border_style, color=color)
    # top
    for row in ws.iter_rows(min_row=min_row,
                            max_row=min_row,
                            min_col=min_col,
                            max_col=max_col):
        for cell in row:
            top = copy(cell.border.top)
            right = copy(cell.border.right)
            bottom = copy(cell.border.bottom)
            left = copy(cell.border.left)
            cell.border = Border(top=border,
                                 right=right,
                                 bottom=bottom,
                                 left=left)
    # right
    for row in ws.iter_rows(min_row=min_row,
                            max_row=max_row,
                            min_col=max_col,
                            max_col=max_col):
        for cell in row:
            top = copy(cell.border.top)
            right = copy(cell.border.right)
            bottom = copy(cell.border.bottom)
            left = copy(cell.border.left)
            cell.border = Border(top=top,
                                 right=border,
                                 bottom=bottom,
                                 left=left)
    # bottom
    for row in ws.iter_rows(min_row=max_row,
                            max_row=max_row,
                            min_col=min_col,
                            max_col=max_col):
        for cell in row:
            top = copy(cell.border.top)
            right = copy(cell.border.right)
            bottom = copy(cell.border.bottom)
            left = copy(cell.border.left)
            cell.border = Border(top=top,
                                 right=right,
                                 bottom=border,
                                 left=left)
    # left
    for row in ws.iter_rows(min_row=min_row,
                            max_row=max_row,
                            min_col=min_col,
                            max_col=min_col):
        for cell in row:
            top = copy(cell.border.top)
            right = copy(cell.border.right)
            bottom = copy(cell.border.bottom)
            left = copy(cell.border.left)
            cell.border = Border(top=top,
                                 right=right,
                                 bottom=bottom,
                                 left=border)


def save_to_excel(rst, xlsx_filename):
    """Pure python codes: save results to excel and as string.
    rst should be sorted.
    """
    # worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = '高性能计算应用中心'
    # headers
    ws.alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=1, column=1, value='姓名')
    ws.cell(row=1, column=2, value='年份季度')
    ws.cell(row=1, column=3, value='手机号')
    ws.cell(row=1, column=4, value='账单月')
    ws.cell(row=1, column=5, value='账单金额')
    ws.cell(row=1, column=6, value='季度/号码')
    ws.cell(row=1, column=7, value='季度/人')
    # fill the sheet
    for i, row in enumerate(rst):
        ws.cell(row=i + 2, column=1, value=rst[i][0])
        ws.cell(row=i + 2, column=2, value=rst[i][1])
        ws.cell(row=i + 2, column=3, value=rst[i][5])
        ws.cell(row=i + 2, column=4, value=rst[i][4])
        ws.cell(row=i + 2, column=5, value=rst[i][6]).number_format = '0.00'
        ws.cell(row=i + 2, column=6).number_format = '0.00'
        ws.cell(row=i + 2, column=7).number_format = '0.00'
    # groupby (u_name, u_phone, y_q_str) and sum
    groups = []
    group_sizes = []
    for k, group in groupby(rst, (0, 1, 5)):
        group = [v[6] for v in group]
        groups.append(group)
        group_sizes.append(len(group))
    group_row_pointers = list(itertools.accumulate([2] + group_sizes))
    # sum and merge
    for i in range(len(groups)):
        # quarter_sum_per_phone
        start_row = group_row_pointers[i]
        end_row = group_row_pointers[i + 1] - 1
        excel_merge_cells(ws,
                          start_row=start_row,
                          end_row=end_row,
                          start_column=6,
                          end_column=6,
                          value=sum(groups[i]))
        # u_phone
        excel_merge_cells(ws,
                          start_row=start_row,
                          end_row=end_row,
                          start_column=3,
                          end_column=3)
    # groupby (u_name, y_q_str) and sum
    groups = []
    group_sizes = []
    for k, group in groupby(rst, (0, 1)):
        group = [v[6] for v in group]
        groups.append(group)
        group_sizes.append(len(group))
    group_row_pointers = list(itertools.accumulate([2] + group_sizes))
    # sum and merge
    for i in range(len(groups)):
        # quarter_sum_per_user
        start_row = group_row_pointers[i]
        end_row = group_row_pointers[i + 1] - 1
        excel_merge_cells(ws,
                          start_row=start_row,
                          end_row=end_row,
                          start_column=7,
                          end_column=7,
                          value=sum(groups[i]))
        # y_q_str
        excel_merge_cells(ws,
                          start_row=start_row,
                          end_row=end_row,
                          start_column=2,
                          end_column=2)
        # border
        set_outline_border(ws,
                           min_row=start_row,
                           max_row=end_row,
                           border_style='thin')
        set_outline_border(ws,
                           min_row=start_row,
                           max_row=end_row,
                           min_col=1,
                           max_col=1,
                           border_style='thin')
    # groupby (u_name) and sum
    groups = []
    group_sizes = []
    for k, group in groupby(rst, (0, )):
        group = list(group)
        groups.append(group)
        group_sizes.append(len(group))
    group_row_pointers = list(itertools.accumulate([2] + group_sizes))
    # merge and border
    for i in range(len(groups)):
        excel_merge_cells(ws,
                          start_row=group_row_pointers[i],
                          end_row=group_row_pointers[i + 1] - 1,
                          start_column=1,
                          end_column=1)
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.row_dimensions[1].height = 32
    # alignment
    for row in ws.iter_rows(max_row=len(rst) + 1, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # alignment
    for row in ws.iter_rows(max_row=len(rst) + 1, min_col=5, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    # alignment and font
    for row in ws.iter_rows(max_row=1, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=11)
    # out line bolder
    set_outline_border(ws, min_row=1, max_row=1, border_style='medium')
    set_outline_border(ws,
                       min_row=1,
                       max_row=len(rst) + 1,
                       border_style='medium')
    wb.save(xlsx_filename)


def print_invoice(input_dir,
                  output_dir,
                  output_filenames,
                  recursive=True,
                  dpi=600,
                  do_analysis=True,
                  poppler_path=None):
    """The main backgroud function that read pdf files from specified folder,
    then do analysis, convert, and save the output.
    """
    outfile_statistics = join(output_dir, output_filenames['stats'])
    outfile_a4_jpg_pdf = join(output_dir, output_filenames['pdf'])
    logger.info('Finding PDF files in directory %r', input_dir)
    pdfs = find_pdfs(input_dir,
                     recursive=recursive,
                     exclude_basenames=output_filenames.values())
    logger.info('PDF files are: %r', pdfs)
    results = []
    if do_analysis is True:
        logger.info('Searching text in the invoice PDF files ...')
        for pdf_path in pdfs:
            text_str = to_text_str(pdf_path, poppler_path=poppler_path)
            results.append(parse_text(text_str) + (pdf_path, ))
        # (user_name, user_phoneno, bill_date, bill_amount, pdf_path)
        logger.info('Aggregating the data ...')
        r_dict = dict(u_names=[],
                      y_q_str=[],
                      b_years=[],
                      b_quarters=[],
                      b_dates=[],
                      u_phones=[],
                      b_amounts=[],
                      pdf_paths=[])
        for u_name, u_phone, b_date, b_amount, pdf_path in results:
            b_year, b_quarter = year_and_quarter(b_date)
            r_dict['u_names'].append(str(u_name))
            r_dict['y_q_str'].append(
                str(b_year) + '年第' + str(b_quarter) + '季度')
            r_dict['b_years'].append(b_year)
            r_dict['b_quarters'].append(b_quarter)
            r_dict['b_dates'].append(str(b_date))
            r_dict['u_phones'].append(str(u_phone))
            r_dict['b_amounts'].append(float(b_amount))
            r_dict['pdf_paths'].append(str(pdf_path))
        results = zip(r_dict['u_names'], r_dict['y_q_str'], r_dict['b_years'],
                      r_dict['b_quarters'], r_dict['b_dates'],
                      r_dict['u_phones'], r_dict['b_amounts'],
                      r_dict['pdf_paths'])
        results = sorted(results, key=operator.itemgetter(0, 1, 5))
        save_to_excel(results, outfile_statistics)
    logger.info('Converting PDF to jpg ...')
    images = pdf_to_jpg(pdfs, dpi=dpi)
    # to_raw_jpg_pdf(images, outfile_raw_jpg_pdf)
    logger.info(
        'Saving images as PDF file with two images on A4-sized page %r',
        outfile_a4_jpg_pdf)
    to_a4_jpg_pdf(images, outfile_a4_jpg_pdf, dpi=dpi)
    logger.info('Backend Done')
